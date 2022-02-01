from flask import redirect, render_template, url_for, request, jsonify, send_from_directory, flash
from flask.helpers import make_response
import os
from app import App
from datetime import datetime, date
from app.models import submitForm, User, Administrator, selectForm
from app.forms import AdminLoginForm
from app.forms import ComplaintForm, COVID19Survey
from app import db
from sqlalchemy import or_

from flask_login import current_user, login_user, login_required, logout_user

from openpyxl import Workbook

def create_path(path):
    if not os.path.exists(path):
        os.makedirs(path)

'''
route for index page
'''
@App.route('/', methods=['GET'])
@App.route('/index', methods=['GET'])
def index():
    return render_template('index.html')

# '''
# route for administrator signup page
# '''
# @app.route('/admin', methods=['GET', 'POST'])
# def login():
#     pass

'''
route for administrator login page
'''
@App.route('/admin', methods=['GET', 'POST'])
def adminlogin():
    if current_user.is_authenticated:
        return render_template('adminview.html', title='Administrator View')
    form = AdminLoginForm()
    if form.validate_on_submit():
        admin = Administrator.query.where(or_(
        Administrator.username==form.emailusername.data, Administrator.email==form.emailusername.data)).first()
        # wrong username or password
        if admin is None or not admin.check_password_hash(form.password.data):
            flash('Invalid username or password.')
            return redirect(url_for('adminlogin'))
        # correct username and password
        login_user(admin, remember = form.remember_me.data)
        # next_page = request.args.get('next')
        # if not next_page or url_parse(next_page).netloc != '':
        #     next_page = url_for('index')
        # return redirect(next_page)
        return render_template('adminview.html', title='Administrator View')
    return render_template('adminlogin.html', title='Administrator Sign In', form=form)

'''
route for administrator logout page
'''
@App.route('/logout', methods=['GET', 'POST'])
@login_required
def adminlogout():
    logout_user()
    return redirect(url_for('adminlogin'))

'''
route for administrative view page
'''
@App.route('/adminview', methods=['GET', 'POST'])
@login_required
def adminview():
    pass

'''
route for forms
formname parameter is the short name of the form
'''
@App.route('/forms/<formname>', methods=['GET', 'POST'])
def renderForm(formname):
    print(formname)
    if formname == "complaintForm":
        form = ComplaintForm()
    elif formname == "COVID19Survey":
        form = COVID19Survey()

    # if request.method == 'POST':
    #     print(f"Username: {form.username.data}\n \
    #     Company name: {form.companyName.data}\n \
    #     Union present: {form.unionPresence.data}\n \
    #     Union head contact no: {form.unionHeadContactNo.data}\n \
    #     Union head email address: {form.unionHeadEmail.data}\n \
    #     Contact number: {form.emailPhone.data}\n \
    #     Complaint: {form.complaint.data}")
    # print(form)
    print(form.validate_on_submit())
    if form.validate_on_submit():
    # if request.method == 'POST':
        # insert the submission to the database
        print('validate')
        username = form.username.data
        contactNumber = form.contactNumber.data
        print(username)
        print(contactNumber)
        questionsAndAnswers = []
        for qa in form:
            # ignore the submit and csrf token fields
            if qa.label.field_id not in ("submit", "csrf_token"):
                questionsAndAnswers.append((qa.label.field_id, qa.data))
        print(questionsAndAnswers)
        submitForm(date.today(), (username, contactNumber), formname, questionsAndAnswers)
        flash("Your form has been submitted. ")
        return redirect(url_for('index'))

    return render_template(f"{formname}.html", title="Complaint Form", form=form)

@App.route('/test')
def test():
    return redirect(url_for('exportFormSubmissions', formname='complaintForm'))

@App.route('/test2')
def test2():
    return redirect(url_for('exportFormSubmissions', formname='COVID19Survey'))


'''
'''
@App.route('/export')
def exportFormSubmissions():
    # get the form name
    formname = request.args.get('formname')

    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("new sheet 1")

    form = selectForm(formname)
    questions = form.questions
    questions.sort(key=lambda q: q.id)
    submissions = form.submissions
    submissions.sort(key=lambda s: s.id)
    print(questions)
    #
    # export general complaint form
    # write header
    question_ids = []
    i=0
    ws['A1'] = 'username'
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=2, max_col=1 + len(questions)):
        for cell in col:
            cell.value = questions[i].short_name
            question_ids.append(questions[i].id)
            i += 1

    print(question_ids)
    # write each submission row
    i = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=1 + len(submissions)):
        submission = submissions[i]
        submission_id = submission.id
        answers = submission.answers
        print(answers)
        username = submission.user.username
        print(username)

        i += 1
        j = 0
        # for each question per row
        for cell in row:
            # write username in first column
            if cell.column == 1:
                cell.value = username
            # other questions
            else:
                question_id = question_ids[j]
                answer = [a for a in answers if a.question_id == question_id and a.submission_id == submission_id]
                print(answer)
                if len(answer):
                    cell.value = answer[0].answer_string
                j += 1
    create_path("files/")
    wb.save("files/x.xlsx")
    print(App.config['UPLOAD_FOLDER'])
    return send_from_directory(App.config['UPLOAD_FOLDER'], 'x.xlsx', as_attachment=True)

@App.route('/uniqueUsername', methods=['GET'])
def checkUsername():
    username = request.form['username']
    unique = User.check_unique_username(username)
    res = 0
    if unique == 1:
        res = 1
    return res
