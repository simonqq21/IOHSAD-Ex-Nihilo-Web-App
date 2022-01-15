from openpyxl import Workbook
# from models import selectForm

wb = Workbook()

ws = wb.active
ws1 = wb.create_sheet("new sheet 1")

# export general complaint form

# write header
i=0
for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=100):
    for cell in col:
        print(cell)
        cell.value = i
        i += 1


wb.save("a.xlsx")
