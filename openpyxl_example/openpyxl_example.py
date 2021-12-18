from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws1 = wb.create_sheet("new sheet 1")
print(ws)
print(ws1)

c = ws['A4']
ws['A4'] = 44

print(c)
print(c.value)

c1 = ws["AA1"]
print(c1)
c1.value = "Hello world openpyxl"

wb.save("example1.xlsx")
